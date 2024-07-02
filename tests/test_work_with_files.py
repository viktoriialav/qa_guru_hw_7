import csv
from io import TextIOWrapper

from openpyxl import load_workbook
from zipfile import ZipFile
from pypdf import PdfReader
import pandas as pd

from set_path import ZIP_FILE


def test_check_xlsx_in_archive():
    file_name = 'FRS.xlsx'
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open(file_name) as file:
            workbook = load_workbook(file)
            title_list = workbook.sheetnames
            sheet = workbook.active
            sheet_2 = workbook[title_list[2]]

            assert sheet.max_column == 8
            assert sheet.max_row == 129
            assert title_list == ['FRS_OX', 'FRS_OY', 'FRS_OZ']
            assert sheet.cell(row=4, column=3).value == 0.112878205
            assert sheet_2.cell(row=10, column=8).value == 0.059621104


def test_check_pdf_in_archive():
    file_name = 'Fluent_Python_2nd_Edition_Luciano_Ramalho.pdf'
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open(file_name) as file:
            reader = PdfReader(file)

            assert len(reader.pages) == 1011
            assert "Now  the focus is on building classes “by hand”" in reader.pages[22].extract_text()
            assert "a Namaqua sand lizard" in reader.pages[1009].extract_text()


def test_check_csv_in_archive_using_csv_module():
    file_name = 'Masses_ISS.csv'
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open(file_name) as file:
            csv_file = file.read().decode(encoding='utf-8')
            content = list(csv.reader(csv_file.splitlines(), delimiter=';'))

            assert len(content), len(content[0]) == (38, 14)
            assert content[16][5] == '1.10E-02'
            assert content[37][13] == '-6.50E-01'


def test_check_csv_in_archive_using_csv_module_and_diff_way_to_read():
    file_name = 'Mass_PCSS.csv'
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open(file_name) as file:
            content = list(csv.reader(TextIOWrapper(file, encoding='utf-8', newline=''), delimiter=';'))

            assert len(content), len(content[0]) == (16, 14)
            assert content[6][4] == '3000'
            assert content[15][13] == '1.70E+00'


def test_check_xlsx_in_archive_using_pandas():
    file_name = 'Building_modal_analysis.xlsx'
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open(file_name) as file_:
            with pd.ExcelFile(file_) as file:
                sheet_1 = pd.read_excel(file, sheet_name=0)
                sheet_2 = pd.read_excel(file, sheet_name=1)
                sheet_3 = pd.read_excel(file, sheet_name=2)

                assert sheet_1.shape == sheet_1.shape == sheet_1.shape
                assert sheet_1.iat[12, 8] == 28150.0
                assert sheet_2.iat[1, 2] == 573.184
                assert sheet_3.iat[16, 6] == 46983.6


def test_check_dat_in_archive():
    file_name = 'ds.dat'
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open(file_name) as file:
            for _ in range(5):
                file.readline().decode('utf-8')

            assert file.readline().decode('utf-8').strip() == '/title,Frame_PCSS--LC#1+0.4X +0.4Y -Z (F3)'

            for _ in range(17):
                file.readline().decode('utf-8')

            assert file.readline().decode('utf-8').strip() == 'nblock,3,,522771'
